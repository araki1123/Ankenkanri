from asyncio.windows_events import NULL
from pickle import NONE
from re import I
from unittest.mock import NonCallableMock
from django.shortcuts import render
from django.http import HttpResponse

import sqlite3
import openpyxl
import os
# from numpy import true_divide
# from django import setup
from .models import AnkenList

# Create your views here.

# def index(request):
#     return HttpResponse('<h1> Good Bye </h1>')

def index (request):
    context = {
        'ankenList' : AnkenList.objects.all()
    }
    return render(request,'index.html',context)

def home(request):
    my_name = 'Taro Yamada'
    favourite_fruits = ['Apple','Grape','Lemon']
    my_info = {
        'name': 'Taro',
        'age': 18
    }
    return render(request, 'home.html',context={
        'my_name':my_name,
        'favourite_fruits':favourite_fruits,
        'my_info':my_info,
    })

def sample (request):
    name = 'ichiro yamada'
    height = 175.5
    weight = 72
    bmi = weight / (height /100)**2
    page_url = 'https://www.google.com'
    favourete_fruits = [
        'Apple','Grape','Lemon'
    ]
    msg = """hello
    my name is
    ichiro
    """
    msg2 = '1234567890'
    return render(request, 'sample.html',context = {
        'name':name,
        'bmi' : bmi,
        'page_url':page_url,
        'fruits':favourete_fruits,
        'msg':msg,
        'msg2':msg2
    })

class Country:
    
    def __init__(self, name, population, capital):
        self.name = name
        self.population = population
        self.capital = capital

def sample3(request):
    country = Country('Japan', 100000000, 'Tokyo')
    return render(request,'sample3.html',context={
        'country':country
    })
    
def export2 (request):
    inputFile=request.POST.get('inputFile')
    context = {
        'ankenList' : AnkenList.objects.all(),
        'inputFile' : inputFile
    }
    if request.method == "POST":
        if "inputFile" in request.POST:
 
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet['A1'].value = 'id'
            sheet['B1'].value = '????????????'
            sheet['C1'].value = '????????????????????????'
            sheet['D1'].value = '?????????????????????'
            sheet['E1'].value = '??????????????????'
            sheet['F1'].value = '??????????????????'
            sheet['G1'].value = '??????No.'
            sheet['H1'].value = '??????'
            sheet['I1'].value = '?????????'
            sheet['J1'].value = '??????????????????'
            sheet['K1'].value = '????????????'
            sheet['L1'].value = '????????????'
            sheet['M1'].value = '????????????A'
            sheet['N1'].value = '?????????B'
            sheet['O1'].value = '???????????????????????????C=A??B'
            sheet['P1'].value = '????????????D'
            sheet['Q1'].value = '?????????E'
            sheet['R1'].value = '???????????????????????????F=D??E'
            sheet['S1'].value = '??????????????????'
            sheet['T1'].value = 'wfNo'
            sheet['U1'].value = '????????????'
            sheet['V1'].value = '????????????G'
            sheet['W1'].value = '?????????H'
            sheet['X1'].value = '???????????????????????????I=G??H'
            sheet['Y1'].value = '??????????????????'
            sheet['Z1'].value = '????????????J'
            sheet['AA1'].value = '?????????K'
            sheet['AB1'].value = '???????????????????????????L=J??K'
            sheet['AC1'].value = '????????????M'
            sheet['AD1'].value = '?????????N'
            sheet['AE1'].value = '???????????????????????????O=M??N'
            sheet['AF1'].value = '??????????????????'
            sheet['AG1'].value = '??????????????????G=C-O'
            sheet['AH1'].value = '???????????????'
            sheet['AI1'].value = '???????????????'
            sheet['AJ1'].value = '???????????????'
            sheet['AK1'].value = '?????????????????????'
            sheet['AL1'].value = '???????????????'
            sheet['AM1'].value = '?????????'
            sheet['AN1'].value = '????????????'
            sheet['AO1'].value = '????????????????????????'
            sheet['AP1'].value = '???????????????'
            sheet['AQ1'].value = '????????????'
            sheet['AR1'].value = '????????????'
            sheet['AS1'].value = '????????????'


            db = sqlite3.connect('db.sqlite3')
            c = db.cursor()
            c.execute('SELECT * FROM ankenListTable')
            for i, row in enumerate (c):
                sheet['A' + str(i+2)].value = row[0]
                sheet['B' + str(i+2)].value = row[1]
                sheet['C' + str(i+2)].value = row[2]
                sheet['D' + str(i+2)].value = row[3]
                sheet['E' + str(i+2)].value = row[4]
                sheet['F' + str(i+2)].value = row[5]
                sheet['G' + str(i+2)].value = row[6]
                sheet['H' + str(i+2)].value = row[7]
                sheet['I' + str(i+2)].value = row[8]
                sheet['J' + str(i+2)].value = row[9]
                sheet['K' + str(i+2)].value = row[10]
                sheet['L' + str(i+2)].value = row[11]
                sheet['M' + str(i+2)].value = row[12]
                sheet['N' + str(i+2)].value = row[13]
                sheet['O' + str(i+2)].value = row[14]
                sheet['P' + str(i+2)].value = row[15]
                sheet['Q' + str(i+2)].value = row[16]
                sheet['R' + str(i+2)].value = row[17]
                sheet['S' + str(i+2)].value = row[18]
                sheet['T' + str(i+2)].value = row[19]
                sheet['U' + str(i+2)].value = row[20]
                sheet['V' + str(i+2)].value = row[21]
                sheet['W' + str(i+2)].value = row[22]
                sheet['X' + str(i+2)].value = row[23]
                sheet['Y' + str(i+2)].value = row[24]
                sheet['Z' + str(i+2)].value = row[25]
                sheet['AA' + str(i+2)].value = row[26]
                sheet['AB' + str(i+2)].value = row[27]
                sheet['AC' + str(i+2)].value = row[28]
                sheet['AD' + str(i+2)].value = row[29]
                sheet['AE' + str(i+2)].value = row[30]
                sheet['AF' + str(i+2)].value = row[31]
                sheet['AG' + str(i+2)].value = row[32]
                sheet['AH' + str(i+2)].value = row[33]
                sheet['AI' + str(i+2)].value = row[34]
                sheet['AJ' + str(i+2)].value = row[35]
                sheet['AK' + str(i+2)].value = row[36]
                sheet['AL' + str(i+2)].value = row[37]
                sheet['AM' + str(i+2)].value = row[38]
                sheet['AN' + str(i+2)].value = row[39]
                sheet['AO' + str(i+2)].value = row[40]
                sheet['AP' + str(i+2)].value = row[41]
                sheet['AQ' + str(i+2)].value = row[42]
                sheet['AR' + str(i+2)].value = row[43]
                sheet['AS' + str(i+2)].value = row[44]
            c.close()
            workbook.save(inputFile) 
 
    
        
    return render(request,'export2.html',context)


        
def dataimport (request):
    context = {
        'ankenList' : AnkenList.objects.all()
    }
    return render(request,'import.html',context)
   
   
def dataexport (request):
    context = {
        'ankenList' : AnkenList.objects.all()
    }
    return render(request,'export.html',context)
     

    
def import2 (request):
    os.environ.setdefault('DJANGO_SETTINGS_MODULE','first_project.settings')
    from django import setup
    setup()
    
    inputFile=request.POST.get('inputFile')
    print('??????????????????????????????',inputFile)
    if request.method == "POST":
        if "inputFile" in request.POST:
 
            workbook = openpyxl.load_workbook(inputFile)
            sheet = workbook.active
            it = sheet.iter_rows(min_row=2)
            kanriNos = []
            edabans = []
            for row in it:
                kanriNos.append(row[6].value)
                edabans.append(row[7].value)
                # print('ID?????????',row[0].value)
                # AnkenList.objects.update_or_create (
                ankn, created=AnkenList.objects.update_or_create (    
                    kanriNo= row[6].value,
                    edaban= row[7].value,
                    defaults={
                        "keiriShonin": row[1].value,              
                        "statusCord": row[2].value,
                        "status":  row[3].value,
                        "edabanGroup":  row[4].value,
                        "shiharaiPattern":  row[5].value,
                        "kanriNo": row[6].value,
                        "edaban": row[7].value,
                        "ankenMei":  row[8].value,
                        "torihikisakiCode":  row[9].value,
                        "torihikisakiMei":  row[10].value,
                        "kanjokamoku":  row[11].value,
                        "keikakuTanka":  row[12].value,
                        "keikakuSuu":  row[13].value,
                        "keikakuGokei":  row[14].value,
                        "mitsumoriTanka":  row[15].value,
                        "mitsumoriSuu":  row[16].value,
                        "mitsumoriGokei":  row[17].value,
                        "mtsumoriLink":  row[18].value,
                        "wfNo":  row[19].value,
                        "keiyakuKingaku":  row[20].value,
                        "chumonTanka":  row[21].value,
                        "chumonSuu":  row[22].value,
                        "chumonGokei":  row[23].value,
                        "chumonLink":  row[24].value,
                        "nohinTanka":  row[25].value,
                        "nohinSuu":  row[26].value,
                        "nohinGokei":  row[27].value,
                        "shiharaiTanka":  row[28].value,
                        "konyuSuu":  row[29].value,
                        "konyuGokei":  row[30].value,
                        "seikyushoLink":  row[31].value,
                        "keikaku_Jisseki":  row[32].value,
                        "nohinKigen":  row[33].value,
                        "kenshuKigen":  row[34].value,
                        "shiharaiKigen":  row[35].value,
                        "kurikaeshiKigen":  row[36].value,
                        "ringishoLink":  row[37].value,
                        "keiyakushoLink":  row[38].value,
                        "konyuKaisha":  row[39].value,
                        "dataKoshinbi":  row[40].value,
                        "saishuKoshinsha":  row[41].value,
                        "shainNo":  row[42].value,
                        "tantosha":  row[43].value,
                        "comment":  row[44].value,
                    }    
                )
                

            # print('??????????????????',ids) 
            ankenList=AnkenList.objects.all()   
            for anken in ankenList:
                hantei=0
                for edabanss in edabans:
                    for kanriNoss in kanriNos:
                        if anken.kanriNo==kanriNoss:
                            if anken.edaban==edabanss:
                                hantei=1
                if hantei==0:
                    AnkenList.objects.filter(id=anken.id).delete()                          

    context = {
        'ankenList' : AnkenList.objects.all(),
        'inputFile' : inputFile
    }            
             
          
    return render(request,'import2.html',context)


def import22 (request):
    os.environ.setdefault('DJANGO_SETTINGS_MODULE','first_project.settings')
    from django import setup
    setup()
    
    inputFile=request.POST.get('inputFile')
    # print('??????????????????????????????',inputFile)
    
    # if request.method == "POST":
        # if "inputFile" in request.POST:
             
    workbook = openpyxl.load_workbook(inputFile)
    sheet = workbook.active
    it = sheet.iter_rows(min_row=2)
    ids = []
    
    for row in it:
        ids.append(row[0].value)
        # a = AnkenList(
        
        #     keiriShonin= True,
        #     statusCord= 0,
        #     status= '???????????????????????????',
        #     edabanGroup= 3,
        #     shiharaiPattern= 3,
        #     kanriNo= 'S22-010',
        #     edaban= '1',
        #     ankenMei= 'Office365 ???????????????4???',
        #     saishuKoshinsha= '??????',
        #     shainNo= 17059,
        #     tantosha= '??????',
        #     comment= '????????????'
        # )

        # a.save(force_insert=True)    
        
        
        # print('ID?????????',row[0].value)
    
        ankn, created=AnkenList.objects.update_or_create (    
            kanriNo= row[6].value,
            edaban= row[7].value,
            defaults={
                "keiriShonin": row[1].value,              
                "statusCord": row[2].value,
                "status":  row[3].value,
                "edabanGroup":  row[4].value,
                "shiharaiPattern":  row[5].value,
                "kanriNo": row[6].value,
                "edaban": row[7].value,
                "ankenMei":  row[8].value,
                "torihikisakiCode":  row[9].value,
                "torihikisakiMei":  row[10].value,
                "kanjokamoku":  row[11].value,
                "keikakuTanka":  row[12].value,
                "keikakuSuu":  row[13].value,
                "keikakuGokei":  row[14].value,
                "mitsumoriTanka":  row[15].value,
                "mitsumoriSuu":  row[16].value,
                "mitsumoriGokei":  row[17].value,
                "mtsumoriLink":  row[18].value,
                "wfNo":  row[19].value,
                "keiyakuKingaku":  row[20].value,
                "chumonTanka":  row[21].value,
                "chumonSuu":  row[22].value,
                "chumonGokei":  row[23].value,
                "chumonLink":  row[24].value,
                "nohinTanka":  row[25].value,
                "nohinSuu":  row[26].value,
                "nohinGokei":  row[27].value,
                "shiharaiTanka":  row[28].value,
                "konyuSuu":  row[29].value,
                "konyuGokei":  row[30].value,
                "seikyushoLink":  row[31].value,
                "keikaku_Jisseki":  row[32].value,
                "nohinKigen":  row[33].value,
                "kenshuKigen":  row[34].value,
                "shiharaiKigen":  row[35].value,
                "kurikaeshiKigen":  row[36].value,
                "ringishoLink":  row[37].value,
                "keiyakushoLink":  row[38].value,
                "konyuKaisha":  row[39].value,
                "dataKoshinbi":  row[40].value,
                "saishuKoshinsha":  row[41].value,
                "shainNo":  row[42].value,
                "tantosha":  row[43].value,
                "comment":  row[44].value,
                
                
                
                
                
                
                
            }    
        )
        
        # if created:
        #     print('???????????????', ankn)
        # else:
        #     print('???????????????', ankn)
        

    ankenList=AnkenList.objects.all()   
    for anken in ankenList:
        hantei=0
        for idss in ids:
            if anken.id==idss:
                hantei=1
        if hantei==0:
            AnkenList.objects.filter(id=anken.id).delete()                          
                    
                
                    

    context = {
        'ankenList' : AnkenList.objects.all(),
        'inputFile' : inputFile
    }            
             
          
    return render(request,'import2.html',context)



def importTest (request):
    os.environ.setdefault('DJANGO_SETTINGS_MODULE','first_project.settings')
    from django import setup
    setup()
    # AnkenList.objects.create(
    a = AnkenList(
        
        keiriShonin= True,
        statusCord= 0,
        status= '???????????????????????????',
        edabanGroup= 3,
        shiharaiPattern= 3,
        kanriNo= 'S22-010',
        edaban= '1',
        ankenMei= 'Office365 ???????????????4???',
        saishuKoshinsha= '??????',
        shainNo= 17059,
        tantosha= '??????',
        comment= '????????????'
    )
    a.save()
    context = {
        'ankenList' : AnkenList.objects.all(),
        
    }
    return render(request,'import2.html',context)